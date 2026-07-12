#!/usr/bin/env python3
"""
#11 wood multi-category overlay + #10 conditioned PAV/SD test.

#11: classify Picea abies wood genes by Rodriguez et al. (2025) clique
     conservation (conserved across 6 species / lineage-differentiated /
     gymnosperm-specific) and overlay our cross-species YN00 dN/dS.
#10: test whether PAV enrichment in the not_coex co-expression category is
     independent of segmental-duplication status.

Runs from any working directory (self-locates the repo root). Writes:
  results/integration/wood_clique_dnds.tsv
  results/integration/pav_conditioned_enrichment.tsv
"""
import os
import openpyxl, pandas as pd, numpy as np
from scipy.stats import kruskal, mannwhitneyu, fisher_exact
import statsmodels.api as sm

# Resolve all relative paths against the repo root (AbioticStressConifers), so the
# script runs correctly regardless of the caller's working directory.
os.chdir(os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..")))

INTEG = "results/integration"
WOOD  = "data/wood"  # Rodriguez et al. (2025) Table S4A (wood_clique_genes.xlsx);
                                           # external, not redistributed - place the extracted file here

# ---------------- #11 wood clique overlay ----------------
def nor(sheet):
    wb = openpyxl.load_workbook(f"{WOOD}/wood_clique_genes.xlsx", read_only=True)
    if sheet not in wb.sheetnames: wb.close(); return set()
    ws = wb[sheet]; hdr=None; g=set()
    for r in ws.iter_rows(values_only=True):
        if hdr is None:
            if r and "OrthoGroup" in [str(x) for x in r]: hdr={str(c):i for i,c in enumerate(r)}
            continue
        i=hdr.get("P. abies")   # Norway spruce gene column in the published Rodriguez et al. (2025) Table S4A
        if i is not None and i<len(r) and r[i] and str(r[i]).startswith("PA_"): g.add(str(r[i]).strip())
    wb.close(); return g

cons = nor("Conserved genes unique")
gym  = nor("Conifer-specific genes unique")   # gymnosperm-specific cliques (sheet name in published S4A)
diff = nor("Differentiated genes")
diff -= cons; gym -= (cons | diff)
wclass = {**{g:"wood_conserved" for g in cons},
          **{g:"wood_differentiated" for g in diff},
          **{g:"wood_gymno_specific" for g in gym}}

bb = pd.read_csv(f"{INTEG}/integration_backbone_1to1.tsv", sep="\t")[["pa_gene","ps_gene"]]
yn = pd.read_csv(f"{INTEG}/cross_species_dnds_yn00.tsv", sep="\t")[["pa_gene","ps_gene","dS","dNdS"]]
# Fig 5a reliability filter (dS in (0,5), dNdS<10) — consistent with figure5_dnds_evolution.R
dn = bb.merge(yn, on=["pa_gene","ps_gene"]); dn = dn[(dn.dS>0) & (dn.dS<5) & (dn.dNdS<10)]
dn["w"] = dn.pa_gene.map(wclass).fillna("background")
order = ["wood_conserved","wood_differentiated","wood_gymno_specific","background"]
rows=[]
for c in order:
    v = dn[dn.w==c]["dNdS"]; rows.append([c,len(v),round(v.median(),4)])
wt = pd.DataFrame(rows, columns=["wood_class","n","median_dNdS"])
H,p = kruskal(*[dn[dn.w==c]["dNdS"].values for c in order[:3]])
wt["KW_H_3class"]=round(H,2); wt["KW_p_3class"]=p
wt["vs_conserved_p"]=[np.nan]+[mannwhitneyu(dn[dn.w=='wood_conserved']['dNdS'],dn[dn.w==c]['dNdS'],alternative='two-sided')[1] for c in order[1:]]
wt.to_csv(f"{INTEG}/wood_clique_dnds.tsv", sep="\t", index=False)
print("#11 wood clique dN/dS:"); print(wt.to_string(index=False))

# All pairwise Mann-Whitney comparisons among the three wood classes (docx-facing:
# differentiated vs conserved, differentiated vs gymno-specific, conserved vs gymno-specific)
_wc = ["wood_conserved","wood_differentiated","wood_gymno_specific"]
_pairs=[]
for i in range(len(_wc)):
    for j in range(i+1,len(_wc)):
        a=dn[dn.w==_wc[i]]["dNdS"]; b=dn[dn.w==_wc[j]]["dNdS"]
        _pairs.append([_wc[i],_wc[j],len(a),len(b),
                       round(a.median(),4),round(b.median(),4),
                       mannwhitneyu(a,b,alternative='two-sided')[1]])
wpair=pd.DataFrame(_pairs,columns=["class_a","class_b","n_a","n_b","median_a","median_b","mannwhitney_p"])
wpair.to_csv(f"{INTEG}/wood_clique_dnds_pairwise.tsv", sep="\t", index=False)
print("#11b wood clique pairwise dN/dS:"); print(wpair.to_string(index=False))

# ---------------- #10 conditioned PAV/SD test ----------------
wp = (pd.read_csv("results/ComPlEx/RData/weighted_gene_pairs.tsv", sep="\t",
                  usecols=["Species1","Species2","best_pval","conserved","cold_specific","drought_specific"])
        .rename(columns={"Species1":"pa_gene","Species2":"ps_gene"}).sort_values("best_pval"))
best = wp.drop_duplicates("pa_gene")
catf = lambda r: "conserved" if r.conserved else "cold_specific" if r.cold_specific else "drought_specific" if r.drought_specific else "multi_tissue"
gene_cat = dict(zip(best.pa_gene, best.apply(catf, axis=1)))
u=set()
for f in ["data/expression/SC_expression.txt","data/expression/SD_expression.txt"]:
    u |= set(x for x in pd.read_csv(f, sep="\t", usecols=[0]).iloc[:,0].astype(str) if x.startswith("PA_"))
pav = set(pd.read_csv("sd_popgen_signals.tsv", sep="\t").query("signal.str.contains('pav')", engine="python")["gene"])
k = pd.read_csv("kaks_results.tsv", sep="\t")
so = set(k[k.category=='spruce_only_SD'].gene1)|set(k[k.category=='spruce_only_SD'].gene2)
sh = set(k[k.category=='shared_SD'].gene1)|set(k[k.category=='shared_SD'].gene2)
df = pd.DataFrame({"g":sorted(u)})
df["not_coex"]=(df.g.map(gene_cat).fillna("not_coex")=="not_coex").astype(int)
df["pav"]=df.g.isin(pav).astype(int); df["spruce_only_SD"]=df.g.isin(so).astype(int); df["shared_SD"]=df.g.isin(sh).astype(int)
def ft(sub):
    a=((sub.pav==1)&(sub.not_coex==1)).sum();b=((sub.pav==1)&(sub.not_coex==0)).sum()
    c=((sub.pav==0)&(sub.not_coex==1)).sum();d=((sub.pav==0)&(sub.not_coex==0)).sum()
    OR,p=fisher_exact([[a,b],[c,d]]); return round(OR,2),p,int(a+b)
res=[["unconditioned (all genes)",*ft(df)],
     ["non-spruce_only_SD only",*ft(df[df.spruce_only_SD==0])],
     ["non-SD only",*ft(df[(df.spruce_only_SD==0)&(df.shared_SD==0)])]]
X=sm.add_constant(df[["pav","spruce_only_SD","shared_SD"]].astype(float))
m=sm.Logit(df.not_coex.astype(float),X).fit(disp=0,method="bfgs")
res.append(["logistic not_coex~pav+SD (pav term)", round(np.exp(m.params['pav']),2), m.pvalues['pav'], int(df.pav.sum())])
ct=pd.DataFrame(res, columns=["test","OR","pvalue","n_pav"])
ct.to_csv(f"{INTEG}/pav_conditioned_enrichment.tsv", sep="\t", index=False)
print("\n#10 conditioned PAV->not_coex:"); print(ct.to_string(index=False))
